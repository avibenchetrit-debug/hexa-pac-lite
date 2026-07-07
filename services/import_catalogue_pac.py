"""Import du catalogue PAC depuis un fichier Excel (format fixe multi-onglets).

Chaque onglet : champs en LIGNES (col. A = libellé), modèles en COLONNES.
Ligne d'en-tête = col. A "Champ" ; ligne séparatrice = col. A contient "source".
Zone du haut = specs (-> description_specs), bloc du bas = 9 champs source.
Ne fait AUCUNE écriture : renvoie la liste de modèles au format catalogue.
"""
from openpyxl import load_workbook

# Libellé source (normalisé) -> clé(s) du modèle catalogue
_SOURCE_MAP = {
    "ref interne": "ref",
    "nom commerciale": "nom",
    "usage": "usage",
    "alimentation": "alim",
    "puissance kw (35°c)": ("puiss35", "puiss_chauf"),
    "etas 35°c (%)": "etas35",
    "etas 55°c (%)": "etas55",
    "prix achat ht": "achat",
    "prix vente ttc": "ttc",
}


def _norm_label(v):
    return " ".join(str(v if v is not None else "").strip().lower().split())


def _to_number(v):
    """virgule -> point, renvoie int si entier, float sinon, None si vide/non numérique."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        n = float(v)
    else:
        s = "".join(ch for ch in str(v) if not ch.isspace()).replace(",", ".")
        if not s:
            return None
        try:
            n = float(s)
        except ValueError:
            return None
    return int(n) if n == int(n) else n


def _cell_text(v):
    """Valeur 'telle quelle' pour affichage : '' si vide, entier sans '.0'."""
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _spec_label(a):
    lbl = str(a).strip()
    return "Technologie" if lbl.lower() == "techno" else lbl


def _parse_sheet(ws, warnings):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []
    header_i = separ_i = None
    for i, row in enumerate(rows):
        a = _norm_label(row[0] if row else "")
        if header_i is None and a == "champ":
            header_i = i
        elif header_i is not None and separ_i is None and "source" in a:
            separ_i = i
            break
    if header_i is None or separ_i is None:
        warnings.append(f"Onglet '{ws.title}': zones 'Champ'/'SOURCE' introuvables -> ignoré.")
        return []

    header = rows[header_i]
    model_cols = [j for j in range(1, len(header)) if _cell_text(header[j])]
    spec_rows = [rows[i] for i in range(header_i + 1, separ_i) if _cell_text(rows[i][0] if rows[i] else "")]
    source = {}
    for i in range(separ_i + 1, len(rows)):
        key = _norm_label(rows[i][0] if rows[i] else "")
        if key in _SOURCE_MAP:
            source[key] = rows[i]

    def cell(label_key, j):
        row = source.get(label_key)
        return row[j] if (row is not None and j < len(row)) else None

    out = []
    for j in model_cols:
        ref = _cell_text(cell("ref interne", j))
        ttc = _to_number(cell("prix vente ttc", j))
        col_name = _cell_text(header[j]) or f"col{j+1}"
        if not ref:
            warnings.append(f"Onglet '{ws.title}', modèle '{col_name}': ref vide -> refusé.")
            continue
        if ttc is None:
            warnings.append(f"Onglet '{ws.title}', modèle '{ref}': 'Prix vente TTC' non numérique -> refusé.")
            continue
        puiss = _to_number(cell("puissance kw (35°c)", j))
        model = {
            "ref": ref,
            "nom": _cell_text(cell("nom commerciale", j)),
            "usage": _cell_text(cell("usage", j)),
            "alim": _cell_text(cell("alimentation", j)),
            "puiss35": puiss,
            "puiss_chauf": puiss,
            "etas35": _to_number(cell("etas 35°c (%)", j)),
            "etas55": _to_number(cell("etas 55°c (%)", j)),
            "achat": _to_number(cell("prix achat ht", j)),
            "ttc": ttc,
            "description_specs": [
                {"champ": _spec_label(r[0]), "valeur": _cell_text(r[j] if j < len(r) else "")}
                for r in spec_rows
            ],
        }
        out.append(model)
    return out


def parse_catalogue_xlsx_report(source):
    """Renvoie (models, warnings). 'source' = chemin ou objet file-like (BytesIO)."""
    wb = load_workbook(source, data_only=True, read_only=True)
    models, warnings = [], []
    for ws in wb.worksheets:
        models.extend(_parse_sheet(ws, warnings))
    seen = {}
    for m in models:
        seen[m["ref"]] = seen.get(m["ref"], 0) + 1
    for ref, n in seen.items():
        if n > 1:
            warnings.append(f"Réf '{ref}' présente {n} fois (doublon inter-onglets).")
    return models, warnings


def parse_catalogue_xlsx(source):
    """Renvoie la liste de modèles au format catalogue (9 champs + description_specs)."""
    models, _ = parse_catalogue_xlsx_report(source)
    return models
